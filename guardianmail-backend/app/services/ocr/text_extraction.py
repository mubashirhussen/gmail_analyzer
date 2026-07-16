"""Text extraction across supported document types.

Strategy per format:
  * image/*  → preprocess → Tesseract (pytesseract)
  * pdf      → PyMuPDF for searchable text; falls back to page-rasterise + OCR
               when a page has < 20 characters (i.e. scanned pages).
  * office   → python-docx / openpyxl; xls read via openpyxl if it converted,
               otherwise skipped with a graceful note.
  * text/csv/json/xml → decoded utf-8 (with replacement).
  * zip      → metadata only (never inflate on the OCR worker).

Returned tuple: (text, confidence, page_count, engines_used).
"""
from __future__ import annotations

import io
import zipfile
from typing import Tuple

from app.core.logging import get_logger
from app.services.ocr.config import DEFAULT_DPI, MAX_PDF_PAGES, OCR_LANGS

_log = get_logger(__name__)


def extract(raw: bytes, mime: str) -> Tuple[str, float, int, list[str]]:
    mime = (mime or "").lower()
    if mime.startswith("image/"):
        return _image(raw)
    if mime == "application/pdf":
        return _pdf(raw)
    if mime in {
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/msword",
    }:
        return _docx(raw)
    if mime in {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    }:
        return _xlsx(raw)
    if mime.startswith("text/") or mime in {"application/json", "application/xml"}:
        return raw.decode("utf-8", errors="replace"), 1.0, 1, ["decode"]
    if mime == "application/zip":
        return _zip_manifest(raw), 1.0, 1, ["zip-manifest"]
    return "", 0.0, 0, []


# ---------------------------------------------------------------- images -----
def _image(raw: bytes) -> Tuple[str, float, int, list[str]]:
    from PIL import Image
    import pytesseract

    from app.services.ocr.preprocess import enhance

    prepped = enhance(raw)
    img = Image.open(io.BytesIO(prepped))
    text = pytesseract.image_to_string(img, lang=OCR_LANGS)
    try:
        data = pytesseract.image_to_data(
            img, lang=OCR_LANGS, output_type=pytesseract.Output.DICT
        )
        confs = [int(c) for c in data.get("conf", []) if str(c).lstrip("-").isdigit() and int(c) >= 0]
        conf = (sum(confs) / len(confs) / 100.0) if confs else 0.5
    except Exception:  # pragma: no cover
        conf = 0.5
    return text, conf, 1, ["tesseract"]


# ------------------------------------------------------------------- pdf -----
def _pdf(raw: bytes) -> Tuple[str, float, int, list[str]]:
    engines = ["pymupdf"]
    text_parts: list[str] = []
    scanned_pages = 0
    try:
        import fitz  # PyMuPDF
    except Exception:  # pragma: no cover
        return _pypdf_fallback(raw)

    with fitz.open(stream=raw, filetype="pdf") as doc:
        pages = min(doc.page_count, MAX_PDF_PAGES)
        for i in range(pages):
            page = doc[i]
            t = page.get_text("text") or ""
            if len(t.strip()) >= 20:
                text_parts.append(t)
                continue
            # scanned page — rasterise and OCR
            scanned_pages += 1
            try:
                pix = page.get_pixmap(dpi=DEFAULT_DPI, alpha=False)
                img_bytes = pix.tobytes("png")
                ocr_text, _, _, _ = _image(img_bytes)
                text_parts.append(ocr_text)
            except Exception as e:
                _log.warning("pdf_ocr_page_failed", page=i, error=str(e))

    if scanned_pages:
        engines.append("tesseract")
    text = "\n".join(text_parts)
    conf = 1.0 if scanned_pages == 0 else max(0.4, 1.0 - scanned_pages / (pages or 1) * 0.4)
    return text, conf, pages, engines


def _pypdf_fallback(raw: bytes) -> Tuple[str, float, int, list[str]]:
    from pypdf import PdfReader
    r = PdfReader(io.BytesIO(raw))
    pages = min(len(r.pages), MAX_PDF_PAGES)
    text = "\n".join((r.pages[i].extract_text() or "") for i in range(pages))
    return text, 0.8, pages, ["pypdf"]


# ------------------------------------------------------------------- docx ----
def _docx(raw: bytes) -> Tuple[str, float, int, list[str]]:
    try:
        from docx import Document as _Docx
    except Exception:  # pragma: no cover
        return "", 0.0, 0, []
    d = _Docx(io.BytesIO(raw))
    parts = [p.text for p in d.paragraphs if p.text]
    for table in d.tables:
        for row in table.rows:
            parts.append("\t".join(cell.text for cell in row.cells))
    return "\n".join(parts), 1.0, 1, ["python-docx"]


# ------------------------------------------------------------------ xlsx -----
def _xlsx(raw: bytes) -> Tuple[str, float, int, list[str]]:
    try:
        from openpyxl import load_workbook
    except Exception:  # pragma: no cover
        return "", 0.0, 0, []
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    lines: list[str] = []
    for name in wb.sheetnames:
        ws = wb[name]
        lines.append(f"# sheet: {name}")
        for row in ws.iter_rows(values_only=True):
            lines.append("\t".join("" if v is None else str(v) for v in row))
    return "\n".join(lines), 1.0, len(wb.sheetnames), ["openpyxl"]


# ------------------------------------------------------------------- zip -----
def _zip_manifest(raw: bytes) -> str:
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as z:
            return "\n".join(
                f"{i.filename}\t{i.file_size}\t{'ENCRYPTED' if i.flag_bits & 0x1 else ''}"
                for i in z.infolist()[:500]
            )
    except Exception as e:  # pragma: no cover
        return f"[zip-error] {e}"
